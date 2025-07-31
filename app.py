from flask import Flask, request, jsonify, render_template, redirect, url_for, session, send_file
import sqlite3
import json
import datetime
import uuid
import os
from functools import wraps
from openpyxl import Workbook
from fpdf import FPDF
import io
from werkzeug.utils import secure_filename

app = Flask(__name__)
app.secret_key = 'SAG_secret_key_2025'

# Ensure database exists function
def ensure_db_exists():
    """Ensure database and tables exist before any operation"""
    try:
        conn = sqlite3.connect('sagapi_database.db')
        cursor = conn.cursor()
        
        # Check if users table exists
        cursor.execute("SELECT name FROM sqlite_master WHERE type='table' AND name='users'")
        if not cursor.fetchone():
            init_db()
        
        conn.close()
    except Exception as e:
        print(f"Database initialization error: {e}")
        init_db()

# Database initialization
def init_db():
    conn = sqlite3.connect('sagapi_database.db')
    cursor = conn.cursor()
    
    # Create api_logs table
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS api_logs (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            timestamp TEXT NOT NULL,
            endpoint TEXT NOT NULL,
            agent_name TEXT,
            site TEXT,
            status TEXT NOT NULL,
            response_code INTEGER NOT NULL,
            request_body TEXT,
            response_body TEXT,
            ip_address TEXT
        )
    ''')
    
    # Create receiving_tbs table
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS receiving_tbs (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            document_no TEXT UNIQUE NOT NULL,
            timestamp TEXT NOT NULL,
            partner_id TEXT NOT NULL,
            journal_id TEXT NOT NULL,
            date_order TEXT NOT NULL,
            officers TEXT NOT NULL,
            keterangan_description TEXT,
            driver_name TEXT NOT NULL,
            vehicle_no TEXT NOT NULL,
            destination_warehouse_id TEXT NOT NULL,
            branch_id TEXT NOT NULL,
            original_payload TEXT
        )
    ''')
    
    # Create order_line table
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS order_line (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            receiving_tbs_id INTEGER NOT NULL,
            product_code TEXT NOT NULL,
            qty_brutto REAL NOT NULL,
            qty_tara REAL NOT NULL,
            qty_netto REAL NOT NULL,
            product_uom TEXT NOT NULL,
            sortation_percent REAL,
            sortation_weight REAL,
            qty_netto2 REAL,
            price_unit REAL NOT NULL,
            product_qty INTEGER NOT NULL,
            incoming_date TEXT NOT NULL,
            outgoing_date TEXT NOT NULL,
            FOREIGN KEY (receiving_tbs_id) REFERENCES receiving_tbs (id)
        )
    ''')
    
    # Create users table for login
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS users (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            username TEXT UNIQUE NOT NULL,
            password TEXT NOT NULL,
            role TEXT DEFAULT 'operator'
        )
    ''')
    
    # Insert default admin user
    cursor.execute('''
        INSERT OR IGNORE INTO users (username, password, role) 
        VALUES (?, ?, ?)
    ''', ('admin', 'SAGsecure#2025', 'admin'))
    
    conn.commit()
    conn.close()

# Log API request/response
def log_api_request(endpoint, status, response_code, request_body=None, response_body=None, agent_name=None, site=None):
    conn = sqlite3.connect('sagapi_database.db')
    cursor = conn.cursor()
    
    timestamp = datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    ip_address = request.remote_addr if request else None
    
    cursor.execute('''
        INSERT INTO api_logs (timestamp, endpoint, agent_name, site, status, response_code, request_body, response_body, ip_address)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
    ''', (timestamp, endpoint, agent_name, site, status, response_code, 
          json.dumps(request_body) if request_body else None,
          json.dumps(response_body) if response_body else None, ip_address))
    
    conn.commit()
    conn.close()

# Generate document number
def generate_document_no():
    now = datetime.datetime.now()
    date_str = now.strftime('%Y/%m/%d')
    
    conn = sqlite3.connect('sagapi_database.db')
    cursor = conn.cursor()
    cursor.execute('SELECT COUNT(*) FROM receiving_tbs WHERE DATE(timestamp) = DATE(?)', (now.strftime('%Y-%m-%d'),))
    count = cursor.fetchone()[0] + 1
    conn.close()
    
    return f"TBS/{date_str}/{count:03d}"

# Login required decorator
def login_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'logged_in' not in session:
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    return decorated_function

# Routes
@app.route('/')
def index():
    # Ensure database exists on first access
    ensure_db_exists()
    return redirect(url_for('dashboard'))

# API Health Check Endpoints for Testing
@app.route('/api/health', methods=['GET'])
def api_health():
    """Simple health check endpoint for API testing"""
    return jsonify({
        "status": "healthy",
        "service": "SAGAPI-Proto",
        "version": "1.0.0",
        "timestamp": datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    }), 200

@app.route('/api/test', methods=['GET', 'POST'])
def api_test():
    """Test endpoint that accepts both GET and POST"""
    method = request.method
    data = request.get_json() if request.is_json else None
    
    response = {
        "status": "success",
        "method": method,
        "message": f"API test successful via {method}",
        "timestamp": datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    }
    
    if data:
        response["received_data"] = data
    
    return jsonify(response), 200

@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        # Ensure database exists before login attempt
        ensure_db_exists()
        
        username = request.form['username']
        password = request.form['password']
        
        conn = sqlite3.connect('sagapi_database.db')
        cursor = conn.cursor()
        cursor.execute('SELECT * FROM users WHERE username = ? AND password = ?', (username, password))
        user = cursor.fetchone()
        conn.close()
        
        if user:
            session['logged_in'] = True
            session['username'] = username
            session['role'] = user[3]
            return redirect(url_for('dashboard'))
        else:
            return render_template('login.html', error='Invalid username or password')
    
    return render_template('login.html')

@app.route('/logout')
def logout():
    session.clear()
    return redirect(url_for('login'))

@app.route('/dashboard')
@login_required
def dashboard():
    conn = sqlite3.connect('sagapi_database.db')
    cursor = conn.cursor()
    
    # Get statistics
    cursor.execute('SELECT COUNT(*) FROM api_logs')
    total_requests = cursor.fetchone()[0]
    
    cursor.execute('SELECT COUNT(*) FROM api_logs WHERE response_code = 200')
    success_requests = cursor.fetchone()[0]
    
    cursor.execute('SELECT COUNT(*) FROM api_logs WHERE response_code != 200')
    failed_requests = cursor.fetchone()[0]
    
    cursor.execute('SELECT COUNT(*) FROM receiving_tbs')
    total_tbs = cursor.fetchone()[0]
    
    # Get recent transactions
    cursor.execute('''
        SELECT r.document_no, r.partner_id, r.timestamp, r.driver_name, 
               COUNT(ol.id) as line_count
        FROM receiving_tbs r
        LEFT JOIN order_line ol ON r.id = ol.receiving_tbs_id
        GROUP BY r.id
        ORDER BY r.timestamp DESC
        LIMIT 10
    ''')
    recent_transactions = cursor.fetchall()
    
    # Get daily stats for chart
    cursor.execute('''
        SELECT DATE(timestamp) as date, COUNT(*) as count,
               SUM(CASE WHEN response_code = 200 THEN 1 ELSE 0 END) as success
        FROM api_logs
        WHERE DATE(timestamp) >= DATE('now', '-7 days')
        GROUP BY DATE(timestamp)
        ORDER BY date
    ''')
    daily_stats = cursor.fetchall()
    
    conn.close()
    
    return render_template('dashboard.html', 
                         total_requests=total_requests,
                         success_requests=success_requests,
                         failed_requests=failed_requests,
                         total_tbs=total_tbs,
                         recent_transactions=recent_transactions,
                         daily_stats=daily_stats)

@app.route('/logs')
@login_required
def logs():
    page = request.args.get('page', 1, type=int)
    per_page = 20
    search = request.args.get('search', '')
    status_filter = request.args.get('status', '')
    
    conn = sqlite3.connect('sagapi_database.db')
    cursor = conn.cursor()
    
    query = 'SELECT * FROM api_logs WHERE 1=1'
    params = []
    
    if search:
        query += ' AND (endpoint LIKE ? OR agent_name LIKE ? OR site LIKE ?)'
        search_param = f'%{search}%'
        params.extend([search_param, search_param, search_param])
    
    if status_filter:
        if status_filter == 'success':
            query += ' AND response_code = 200'
        elif status_filter == 'error':
            query += ' AND response_code != 200'
    
    query += ' ORDER BY timestamp DESC'
    
    cursor.execute(query, params)
    all_logs = cursor.fetchall()
    
    total = len(all_logs)
    start = (page - 1) * per_page
    end = start + per_page
    logs_page = all_logs[start:end]
    
    conn.close()
    
    return render_template('logs.html', 
                         logs=logs_page,
                         page=page,
                         total=total,
                         per_page=per_page,
                         search=search,
                         status_filter=status_filter)

@app.route('/log/<int:log_id>')
@login_required
def log_detail(log_id):
    conn = sqlite3.connect('sagapi_database.db')
    cursor = conn.cursor()
    cursor.execute('SELECT * FROM api_logs WHERE id = ?', (log_id,))
    log = cursor.fetchone()
    conn.close()
    
    if not log:
        return "Log not found", 404
    
    return render_template('log_detail.html', log=log)

@app.route('/transactions')
@login_required
def transactions():
    conn = sqlite3.connect('sagapi_database.db')
    cursor = conn.cursor()
    
    cursor.execute('''
        SELECT r.*, 
               COALESCE(COUNT(ol.id), 0) as line_count,
               COALESCE(SUM(ol.qty_netto2), 0) as total_netto,
               COALESCE(SUM(ol.qty_netto2 * ol.price_unit), 0) as total_value
        FROM receiving_tbs r
        LEFT JOIN order_line ol ON r.id = ol.receiving_tbs_id
        GROUP BY r.id
        ORDER BY r.timestamp DESC
    ''')
    raw_transactions = cursor.fetchall()
    
    # Convert to list and ensure numeric values
    transactions = []
    for t in raw_transactions:
        transaction = list(t)
        # Make sure line_count (index 13), total_netto (index 14), total_value (index 15) are numeric
        transaction[13] = int(transaction[13]) if transaction[13] is not None else 0
        transaction[14] = float(transaction[14]) if transaction[14] is not None else 0.0
        transaction[15] = float(transaction[15]) if transaction[15] is not None else 0.0
        transactions.append(tuple(transaction))
    
    conn.close()
    
    return render_template('transactions.html', transactions=transactions)

@app.route('/transaction/<int:transaction_id>')
@login_required
def transaction_detail(transaction_id):
    conn = sqlite3.connect('sagapi_database.db')
    cursor = conn.cursor()
    
    # Get transaction header
    cursor.execute('''
        SELECT * FROM receiving_tbs WHERE id = ?
    ''', (transaction_id,))
    transaction = cursor.fetchone()
    
    if not transaction:
        return "Transaction not found", 404
    
    # Get order lines
    cursor.execute('''
        SELECT * FROM order_line WHERE receiving_tbs_id = ?
        ORDER BY id
    ''', (transaction_id,))
    order_lines = cursor.fetchall()
    
    conn.close()
    
    return render_template('transaction_detail.html', 
                         transaction=transaction, 
                         order_lines=order_lines)

# API Endpoints
@app.route('/api/auth/login', methods=['POST'])
def api_login():
    try:
        data = request.get_json()
        
        if not data:
            response = {"code": 400, "message": "Invalid JSON payload"}
            log_api_request('/api/auth/login', 'error', 400, data, response)
            return jsonify(response), 400
        
        login = data.get('login')
        password = data.get('password')
        database = data.get('database')
        
        if not all([login, password, database]):
            response = {"code": 400, "message": "Missing required fields: login, password, database"}
            log_api_request('/api/auth/login', 'error', 400, data, response)
            return jsonify(response), 400
        
        # Check database
        if database != 'sag_production':
            response = {"code": 400, "message": "Invalid database: database tidak ditemukan."}
            log_api_request('/api/auth/login', 'error', 400, data, response)
            return jsonify(response), 400
        
        # Check credentials
        if login != 'admin' or password != 'SAGsecure#2025':
            response = {"code": 400, "message": "Invalid Login: Incorrect user or password."}
            log_api_request('/api/auth/login', 'error', 400, data, response)
            return jsonify(response), 400
        
        # Generate tokens
        access_token = str(uuid.uuid4())
        refresh_token = str(uuid.uuid4())
        
        response = {
            "code": 200,
            "message": "Login Successful",
            "token": {
                "access_token": access_token,
                "refresh_token": refresh_token
            }
        }
        
        log_api_request('/api/auth/login', 'success', 200, data, response, login, database)
        return jsonify(response), 200
        
    except Exception as e:
        response = {"code": 500, "message": f"Internal server error: {str(e)}"}
        log_api_request('/api/auth/login', 'error', 500, request.get_json(), response)
        return jsonify(response), 500

@app.route('/api/receiving-tbs/create', methods=['POST'])
def api_create_receiving_tbs():
    try:
        data = request.get_json()
        
        if not data:
            response = {"code": 400, "message": "Invalid JSON payload"}
            log_api_request('/api/receiving-tbs/create', 'error', 400, data, response)
            return jsonify(response), 400
        
        # Check authorization - either from header or body
        auth_header = request.headers.get('Authorization')
        token_from_body = data.get('token')
        
        if not auth_header and not token_from_body:
            response = {"code": 401, "message": "Authorization header or token in body required"}
            log_api_request('/api/receiving-tbs/create', 'error', 401, data, response)
            return jsonify(response), 401
        
        # Extract order data
        params = data.get('params', {})
        order_data = params.get('order_data', [])
        
        if not order_data:
            response = {"code": 400, "message": "order_data is required"}
            log_api_request('/api/receiving-tbs/create', 'error', 400, data, response)
            return jsonify(response), 400
        
        order = order_data[0]  # Take first order
        
        # Validate required fields
        required_fields = ['partner_id', 'journal_id', 'date_order', 'officers', 
                          'driver_name', 'vehicle_no', 'destination_warehouse_id', 'branch_id']
        
        for field in required_fields:
            if not order.get(field):
                if field == 'driver_name':
                    response = {"code": 400, "message": "Missing Driver Name. Please fill in this information."}
                elif field == 'partner_id':
                    response = {"code": 400, "message": "Sorry, we couldn't find a valid partner with the information provided."}
                else:
                    response = {"code": 400, "message": f"Missing required field: {field}"}
                log_api_request('/api/receiving-tbs/create', 'error', 400, data, response)
                return jsonify(response), 400
        
        # Validate order lines
        order_lines = order.get('order_line', [])
        if not order_lines:
            response = {"code": 400, "message": "At least one order line is required"}
            log_api_request('/api/receiving-tbs/create', 'error', 400, data, response)
            return jsonify(response), 400
        
        for line in order_lines:
            if not line.get('incoming_date') or not line.get('outgoing_date'):
                response = {"code": 400, "message": "Please provide the correct incoming date and outgoing date."}
                log_api_request('/api/receiving-tbs/create', 'error', 400, data, response)
                return jsonify(response), 400
        
        # Save to database
        conn = sqlite3.connect('sagapi_database.db')
        cursor = conn.cursor()
        
        document_no = generate_document_no()
        timestamp = datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        
        # Insert receiving_tbs
        cursor.execute('''
            INSERT INTO receiving_tbs 
            (document_no, timestamp, partner_id, journal_id, date_order, officers, 
             keterangan_description, driver_name, vehicle_no, destination_warehouse_id, branch_id, original_payload)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        ''', (document_no, timestamp, order['partner_id'], order['journal_id'],
              order['date_order'], order['officers'], order.get('keterangan_description', ''),
              order['driver_name'], order['vehicle_no'], order['destination_warehouse_id'],
              order['branch_id'], json.dumps(data, indent=2)))
        
        receiving_tbs_id = cursor.lastrowid
        
        # Insert order lines
        for line in order_lines:
            cursor.execute('''
                INSERT INTO order_line 
                (receiving_tbs_id, product_code, qty_brutto, qty_tara, qty_netto, product_uom,
                 sortation_percent, sortation_weight, qty_netto2, price_unit, product_qty,
                 incoming_date, outgoing_date)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            ''', (receiving_tbs_id, line['product_code'], line['qty_brutto'], line['qty_tara'],
                  line['qty_netto'], line['product_uom'], line.get('sortation_percent', 0),
                  line.get('sortation_weight', 0), line.get('qty_netto2', line['qty_netto']),
                  line['price_unit'], line['product_qty'], line['incoming_date'], line['outgoing_date']))
        
        conn.commit()
        conn.close()
        
        response = {
            "code": 200,
            "message": "Receiving TBS created successfully.",
            "result": {
                "document_no": document_no
            }
        }
        
        log_api_request('/api/receiving-tbs/create', 'success', 200, data, response, 
                       order.get('partner_id'), order.get('branch_id'))
        return jsonify(response), 200
        
    except Exception as e:
        response = {"code": 500, "message": f"Internal server error: {str(e)}"}
        log_api_request('/api/receiving-tbs/create', 'error', 500, request.get_json(), response)
        return jsonify(response), 500

# Export functions
@app.route('/export/logs/excel')
@login_required
def export_logs_excel():
    conn = sqlite3.connect('sagapi_database.db')
    cursor = conn.cursor()
    cursor.execute('SELECT * FROM api_logs ORDER BY timestamp DESC')
    logs = cursor.fetchall()
    
    # Get column names
    cursor.execute('PRAGMA table_info(api_logs)')
    columns = [column[1] for column in cursor.fetchall()]
    conn.close()
    
    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "API Logs"
    
    # Add headers
    for col, header in enumerate(columns, 1):
        ws.cell(row=1, column=col, value=header)
    
    # Add data
    for row, data in enumerate(logs, 2):
        for col, value in enumerate(data, 1):
            ws.cell(row=row, column=col, value=value)
    
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return send_file(
        output,
        as_attachment=True,
        download_name=f'api_logs_{datetime.datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx',
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )

@app.route('/export/transactions/excel')
@login_required
def export_transactions_excel():
    conn = sqlite3.connect('sagapi_database.db')
    cursor = conn.cursor()
    
    # Get receiving_tbs data
    cursor.execute('SELECT * FROM receiving_tbs ORDER BY timestamp DESC')
    tbs_data = cursor.fetchall()
    cursor.execute('PRAGMA table_info(receiving_tbs)')
    tbs_columns = [column[1] for column in cursor.fetchall()]
    
    # Get order_line data
    cursor.execute('SELECT * FROM order_line ORDER BY receiving_tbs_id')
    lines_data = cursor.fetchall()
    cursor.execute('PRAGMA table_info(order_line)')
    lines_columns = [column[1] for column in cursor.fetchall()]
    
    conn.close()
    
    # Create workbook
    wb = Workbook()
    
    # Create TBS sheet
    ws1 = wb.active
    ws1.title = "Receiving TBS"
    
    # Add TBS headers
    for col, header in enumerate(tbs_columns, 1):
        ws1.cell(row=1, column=col, value=header)
    
    # Add TBS data
    for row, data in enumerate(tbs_data, 2):
        for col, value in enumerate(data, 1):
            ws1.cell(row=row, column=col, value=value)
    
    # Create Order Lines sheet
    ws2 = wb.create_sheet("Order Lines")
    
    # Add Order Lines headers
    for col, header in enumerate(lines_columns, 1):
        ws2.cell(row=1, column=col, value=header)
    
    # Add Order Lines data
    for row, data in enumerate(lines_data, 2):
        for col, value in enumerate(data, 1):
            ws2.cell(row=row, column=col, value=value)
    
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return send_file(
        output,
        as_attachment=True,
        download_name=f'transactions_{datetime.datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx',
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )

if __name__ == '__main__':
    init_db()
    port = int(os.environ.get('PORT', 5000))
    app.run(debug=False, host='0.0.0.0', port=port)
else:
    # For production deployment (gunicorn), ensure database is initialized
    init_db()
